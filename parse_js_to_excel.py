from playwright.sync_api import sync_playwright
from openpyxl import Workbook
import openpyxl

def parse_and_save():
    # 1. Запускаем Playwright
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)  # Меняем на False, чтобы видеть окно браузера
        page = browser.new_page()
        
        try:
            # Увеличиваем timeout до 60 секунд
            page.goto("https://quotes.toscrape.com/js/", timeout=60000)
            
            # Ждем, пока JS отработает и цитаты появятся
            page.wait_for_selector(".quote", timeout=10000)
            
            # Собираем данные
            quotes_data = []
            quote_elements = page.query_selector_all(".quote")
            print(f"Найдено цитат для парсинга: {len(quote_elements)}")
        
            for element in quote_elements:
                text_el = element.query_selector(".text")
                text = text_el.inner_text() if text_el else ""
                
                author_el = element.query_selector(".author")
                author = author_el.inner_text() if author_el else ""
                
                text = text.strip("«»\"""")
                quotes_data.append({"quote": text, "author": author})
            
            print(f"Собрано {len(quotes_data)} цитат")
            
        except Exception as e:
            print(f"Ошибка при парсинге: {e}")
            browser.close()
            return
        
        browser.close()

    # 2. Сохраняем в Excel
    if not quotes_data:
        print("Нет данных для сохранения!")
        return
        
    wb = Workbook()
    ws = wb.active
    ws.title = "Цитаты"
    
    ws.append(["Цитата", "Автор"])
    
    for item in quotes_data:
        ws.append([item["quote"], item["author"]])
        
    # Автоподбор ширины колонок
    for col in ws.columns:
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
    
    # === ФОРМАТИРОВАНИЕ EXCEL ===
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    
    ws.row_dimensions[1].height = 25
    ws.freeze_panes = "A2"
    
    wb.save("quotes.xlsx")
    print("Готово! Файл quotes.xlsx создан.")

if __name__ == "__main__":
    parse_and_save()