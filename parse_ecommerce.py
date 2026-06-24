from playwright.sync_api import sync_playwright
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import random
import time

USER_AGENTS = [
    # Chrome на Windows
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.0.0 Safari/537.36",
    
    # Chrome на Mac
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36",
    
    # Firefox на Windows
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:121.0) Gecko/20100101 Firefox/121.0",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:120.0) Gecko/20100101 Firefox/120.0",
    
    # Firefox на Mac
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10.15; rv:121.0) Gecko/20100101 Firefox/121.0",
    
    # Safari на Mac
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.2 Safari/605.1.15",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.1 Safari/605.1.15",
    
    # Edge на Windows
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36 Edg/120.0.0.0",
    
    # Chrome на Linux
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    
    # Яндекс.Браузер
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 YaBrowser/24.1.0.0 Safari/537.36",
]

def human_delay(min_sec=1, max_sec=3):
    time.sleep(random.uniform(min_sec, max_sec))

def parse_labirint():
    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=False,
            args=['--disable-blink-features=AutomationControlled']
        )
        context = browser.new_context(
            user_agent=random.choice(USER_AGENTS),
            viewport={'width': 1920, 'height': 1080},
        )
        page = context.new_page()
        
        books_data = []
        base_url = "https://www.labirint.ru/genres/2308/"
        
        try:
            for page_num in range(1, 4):
                if page_num == 1:
                    url = base_url
                else:
                    url = f"{base_url}?page={page_num}"
                
                print(f"\n=== Страница {page_num}: {url} ===")
                
                try:
                    page.goto(url, timeout=60000)
                    page.wait_for_timeout(3000)
                except Exception as e:
                    print(f"⚠️  Не удалось загрузить страницу {page_num}: {e}")
                    break
                
                # Ищем ссылки на книги (это и есть карточки)
                book_links = page.locator("a.product-title-link")
                count = book_links.count()
                print(f"Найдено книг: {count}")
                
                # Ищем все цены
                prices = page.locator(".price-val")
                prices_count = prices.count()
                print(f"Найдено цен: {prices_count}")
                
                # Собираем данные
                for i in range(count):
                    link_el = book_links.nth(i)
                    
                    # Ссылка
                    href = link_el.get_attribute("href")
                    full_link = f"https://www.labirint.ru{href}" if href and href.startswith('/') else href
                    
                    # Название (внутри ссылки)
                    title_el = link_el.locator(".product-title")
                    if title_el.count() > 0:
                        title = title_el.inner_text()
                    else:
                        title = link_el.inner_text()
                    
                    # Цена (сопоставляем по индексу)
                    price = ""
                    if i < prices_count:
                        price = prices.nth(i).inner_text()
                    
                    books_data.append({
                        "title": title.strip(),
                        "price": price.strip() if price else "0",
                        "link": full_link,
                        "page": page_num
                    })
                    
                    if i < 5:
                        print(f"  {i+1}. {title[:50]}... - {price} руб.")
                        print(f"     Ссылка: {full_link}")
                
                if page_num < 3:
                    print(f"  Ждём перед следующей страницей...")
                    human_delay(2, 4)
                    
        except Exception as e:
            print(f"Ошибка: {e}")
            import traceback
            traceback.print_exc()
        
        browser.close()

    # Сохраняем в Excel
    if not books_data:
        print("\nНет данных для сохранения!")
        return
        
    wb = Workbook()
    ws = wb.active
    ws.title = "Книги"
    
    ws.append(["Название", "Цена (руб.)", "Ссылка", "Страница"])
    
    for book in books_data:
        ws.append([book["title"], book["price"], book["link"], book["page"]])
        
    # Автоширина колонок
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col_letter].width = min(max_len + 5, 80)
        
    # Форматирование шапки
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        
    wb.save("labirint_books.xlsx")
    print(f"\n✅ Готово! Собрано {len(books_data)} книг в labirint_books.xlsx")

if __name__ == "__main__":
    parse_labirint()